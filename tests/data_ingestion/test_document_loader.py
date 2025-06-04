import pytest
import asyncio
import os
from unittest import mock
from pathlib import Path # Import Path
import logging

# Make sure aiofiles is available for mocking
try:
    import aiofiles
    import aiofiles.os as aio_os # Though not directly used by current loader
except ImportError:
    aiofiles = mock.MagicMock()
    aio_os = mock.MagicMock()

# Import the class to be tested
from data_ingestion.document_loader import DocumentLoader

# Actual text extensions supported by _load_text_async in DocumentLoader
TEXT_EXTENSIONS = ['.txt', '.md', '.py', '.js', '.java', '.html', '.css', '.json', '.yaml', '.log']
# Other directly supported extensions
OTHER_EXTENSIONS = ['.pdf', '.docx', '.xlsx']
ALL_SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS + OTHER_EXTENSIONS

@pytest.fixture
def doc_loader():
    # DocumentLoader has only static methods, so an instance isn't strictly necessary
    # but if it were to have instance state later, this would be useful.
    return DocumentLoader()

# --- Mocks for synchronous loaders ---
@pytest.fixture
def mock_load_pdf_sync():
    with mock.patch.object(DocumentLoader, '_load_pdf_sync', return_value="PDF Content") as m:
        yield m

@pytest.fixture
def mock_load_docx_sync():
    with mock.patch.object(DocumentLoader, '_load_docx_sync', return_value="DOCX Content") as m:
        yield m

@pytest.fixture
def mock_load_xlsx_sync():
    with mock.patch.object(DocumentLoader, '_load_xlsx_sync', return_value="XLSX Content") as m:
        yield m

@pytest.fixture
def mock_load_text_async():
    # This mocks the internal _load_text_async static method
    with mock.patch.object(DocumentLoader, '_load_text_async', return_value="Text Content") as m:
        yield m

@pytest.fixture
def mock_asyncio_to_thread():
    # To mock asyncio.to_thread used for sync loaders
    # It needs to return the result of the function it's supposed to run
    async def passthrough_to_thread(func, *args):
        return func(*args) # Directly execute the sync function for testing ease

    with mock.patch('asyncio.to_thread', side_effect=passthrough_to_thread) as m:
        yield m

# --- Tests for load_document ---

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True) # Assume file exists for these tests
async def test_load_document_pdf(mock_os_exists, doc_loader, mock_asyncio_to_thread, mock_load_pdf_sync):
    filepath = "test.pdf"
    content = await doc_loader.load_document(filepath)
    assert content == "PDF Content"
    mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_pdf_sync, filepath)
    mock_load_pdf_sync.assert_called_once_with(filepath)

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_docx(mock_os_exists, doc_loader, mock_asyncio_to_thread, mock_load_docx_sync):
    filepath = "test.docx"
    content = await doc_loader.load_document(filepath)
    assert content == "DOCX Content"
    mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_docx_sync, filepath)
    mock_load_docx_sync.assert_called_once_with(filepath)

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_xlsx(mock_os_exists, doc_loader, mock_asyncio_to_thread, mock_load_xlsx_sync):
    filepath = "test.xlsx"
    content = await doc_loader.load_document(filepath)
    assert content == "XLSX Content"
    mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_xlsx_sync, filepath)
    mock_load_xlsx_sync.assert_called_once_with(filepath)

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_txt(mock_os_exists, doc_loader, mock_load_text_async):
    # Test one of the text extensions
    filepath = "test.txt"
    content = await doc_loader.load_document(filepath)
    assert content == "Text Content"
    mock_load_text_async.assert_called_once_with(filepath)

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_py(mock_os_exists, doc_loader, mock_load_text_async):
    # Test another text_extension
    filepath = "script.py"
    content = await doc_loader.load_document(filepath)
    assert content == "Text Content"
    mock_load_text_async.assert_called_once_with(filepath)


@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=False) # File does NOT exist
async def test_load_document_non_existent_file(mock_os_exists, doc_loader, caplog):
    filepath = "non_existent.txt"
    caplog.set_level(logging.ERROR)
    content = await doc_loader.load_document(filepath)
    assert content is None
    assert f"File not found: {filepath}" in caplog.text
    mock_os_exists.assert_called_once_with(filepath)

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_unsupported_extension(mock_os_exists, doc_loader, caplog):
    filepath = "test.unsupported"
    caplog.set_level(logging.WARNING)
    content = await doc_loader.load_document(filepath)
    assert content is None
    assert f"Unsupported file type: .unsupported for file {filepath}" in caplog.text

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_read_error_pdf(mock_os_exists, doc_loader, mock_asyncio_to_thread, mock_load_pdf_sync, caplog):
    filepath = "error.pdf"
    mock_load_pdf_sync.side_effect = Exception("PDF Read Error")
    caplog.set_level(logging.ERROR)

    content = await doc_loader.load_document(filepath)

    assert content is None
    assert f"Error loading document {filepath}: PDF Read Error" in caplog.text
    mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_pdf_sync, filepath)

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_read_error_text(mock_os_exists, doc_loader, mock_load_text_async, caplog):
    filepath = "error.txt"
    mock_load_text_async.side_effect = Exception("Text Read Error")
    caplog.set_level(logging.ERROR)

    content = await doc_loader.load_document(filepath)
    assert content is None
    assert f"Error loading document {filepath}: Text Read Error" in caplog.text

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_empty_file_text(mock_os_exists, doc_loader, mock_load_text_async):
    filepath = "empty.txt"
    mock_load_text_async.return_value = "" # Simulate empty content
    content = await doc_loader.load_document(filepath)
    assert content == "" # Empty string is valid content


# --- Tests for _load_text_async internal method ---
# These test the fallback logic within _load_text_async specifically

@pytest.mark.asyncio
async def test_internal_load_text_async_success_aiofiles(doc_loader):
    """ Test _load_text_async succeeds with aiofiles """
    filepath = "test.txt"
    mock_content = "aiofiles success"

    # Correctly configure the async context manager mock for aiofiles.open
    async_file_mock = mock.AsyncMock() # This is the object the 'async with' yields (the file handle)
    async_file_mock.read = mock.AsyncMock(return_value=mock_content) # Configure its read method

    # mock_aiofiles_open_cm is what aiofiles.open(...) itself returns.
    # This object needs to be an async context manager.
    mock_aiofiles_open_cm = mock.AsyncMock()
    mock_aiofiles_open_cm.__aenter__.return_value = async_file_mock # __aenter__ returns the file handle mock

    with mock.patch("aiofiles.open", return_value=mock_aiofiles_open_cm) as mocked_aio_open_patch:
        with mock.patch("builtins.open") as mocked_sync_open: # Should not be called
            content = await DocumentLoader._load_text_async(filepath)
            assert content == mock_content
            mocked_aio_open_patch.assert_called_once_with(filepath, 'r', encoding='utf-8', errors='ignore')
            mocked_sync_open.assert_not_called()

@pytest.mark.asyncio
async def test_internal_load_text_async_aiofiles_fails_fallback_succeeds(doc_loader, caplog):
    """ Test _load_text_async falls back to sync open if aiofiles fails """
    filepath = "test_fallback.txt"
    mock_content_sync = "sync fallback success"

    # Mock aiofiles.open to raise an error
    with mock.patch("aiofiles.open", side_effect=Exception("aiofiles failed")) as mocked_aio_open:
        # Mock builtins.open (for sync fallback)
        mock_sync_file = mock.MagicMock()
        mock_sync_file.read.return_value = mock_content_sync
        mock_sync_open = mock.MagicMock()
        mock_sync_open.return_value.__enter__.return_value = mock_sync_file

        with mock.patch("builtins.open", mock_sync_open) as mocked_sync_open_call:
            caplog.set_level(logging.ERROR)
            content = await DocumentLoader._load_text_async(filepath)

            assert content == mock_content_sync
            mocked_aio_open.assert_called_once_with(filepath, 'r', encoding='utf-8', errors='ignore')
            mocked_sync_open_call.assert_called_once_with(filepath, 'r', encoding='utf-8', errors='ignore')
            assert f"Error reading text file {filepath} asynchronously: aiofiles failed" in caplog.text

@pytest.mark.asyncio
async def test_internal_load_text_async_both_fail(doc_loader, caplog):
    """ Test _load_text_async re-raises error if both aiofiles and sync fallback fail """
    filepath = "test_both_fail.txt"

    with mock.patch("aiofiles.open", side_effect=Exception("aiofiles failed")) as mocked_aio_open:
        with mock.patch("builtins.open", side_effect=Exception("sync failed")) as mocked_sync_open:
            caplog.set_level(logging.ERROR)
            with pytest.raises(Exception, match="sync failed"): # Expecting the sync error to be re-raised
                await DocumentLoader._load_text_async(filepath)

            assert f"Error reading text file {filepath} asynchronously: aiofiles failed" in caplog.text
            assert f"Fallback synchronous read for {filepath} also failed: sync failed" in caplog.text

# --- Tests for load_from_directory ---

@pytest.fixture
def mock_path_rglob():
    # Default rglob mock, can be customized per test
    with mock.patch.object(Path, "rglob") as m:
        yield m

@pytest.fixture
def mock_path_is_dir():
    # Default is_dir mock
    with mock.patch.object(Path, "is_dir") as m:
        yield m

@pytest.mark.asyncio
async def test_load_from_directory_success(doc_loader, mock_path_is_dir, mock_path_rglob, monkeypatch, caplog):
    mock_path_is_dir.return_value = True # Simulate directory exists

    # Simulate file structure from rglob
    # Note: these need to be Path objects
    mock_files = [
        Path("/testdir/file1.txt"),
        Path("/testdir/doc.pdf"),
        Path("/testdir/subdir/image.jpg"), # Unsupported
        Path("/testdir/subdir/archive.zip") # Unsupported
    ]
    # rglob needs to return an iterable of Path objects
    # Each object needs an is_file() method.
    configured_mock_files = []
    for p_obj in mock_files:
        m_file = mock.MagicMock(spec=Path)
        m_file.is_file.return_value = True # All are files
        m_file.__str__.return_value = str(p_obj) # So str(file_path_obj) works
        # Copy over other necessary attributes if Path methods are called on these directly
        m_file.name = p_obj.name
        m_file.parent = p_obj.parent
        configured_mock_files.append(m_file)

    mock_path_rglob.return_value = configured_mock_files

    # Mock DocumentLoader.load_document to check which files it's called with
    # and return predefined content or None for unsupported/error cases
    async def mock_load_document_logic(filepath_str):
        if filepath_str == "/testdir/file1.txt": return "Text content"
        if filepath_str == "/testdir/doc.pdf": return "PDF content"
        if filepath_str == "/testdir/subdir/image.jpg": return None # Skipped by load_document
        if filepath_str == "/testdir/subdir/archive.zip": return None # Skipped
        return None # Default for any other unexpected calls

    # We need to patch the static method on the class for load_from_directory to use the mock
    with mock.patch.object(DocumentLoader, 'load_document', side_effect=mock_load_document_logic) as mock_ld_method:
        caplog.set_level(logging.INFO)
        results = await doc_loader.load_from_directory("/testdir")

    mock_path_is_dir.assert_called_once() # Path("/testdir").is_dir()
    mock_path_rglob.assert_called_once_with('*')

    assert len(results) == 2
    # Sort for stable comparison
    results.sort(key=lambda x: x[0])
    assert results[0] == ("/testdir/doc.pdf", "PDF content")
    assert results[1] == ("/testdir/file1.txt", "Text content")

    # Verify load_document was called for all files found by rglob
    expected_load_doc_calls = [
        mock.call("/testdir/file1.txt"),
        mock.call("/testdir/doc.pdf"),
        mock.call("/testdir/subdir/image.jpg"),
        mock.call("/testdir/subdir/archive.zip"),
    ]
    # Check calls irrespective of order
    assert len(mock_ld_method.call_args_list) == len(expected_load_doc_calls)
    for c in expected_load_doc_calls:
        assert c in mock_ld_method.call_args_list, f"Expected call {c} not found in actual calls."
    # Ensure all actual calls were expected (handles duplicates if any, though not in this case)
    for c_actual in mock_ld_method.call_args_list:
        assert c_actual in expected_load_doc_calls, f"Actual call {c_actual} was not expected."

    assert "Loaded 2 documents from directory /testdir" in caplog.text


@pytest.mark.asyncio
async def test_load_from_directory_not_a_dir(doc_loader, mock_path_is_dir, caplog):
    mock_path_is_dir.return_value = False # Simulate path is not a directory
    caplog.set_level(logging.ERROR)

    results = await doc_loader.load_from_directory("/notadir")

    assert len(results) == 0
    assert "Provided path is not a directory: /notadir" in caplog.text

@pytest.mark.asyncio
async def test_load_from_directory_one_file_fails(doc_loader, mock_path_is_dir, mock_path_rglob, caplog):
    mock_path_is_dir.return_value = True

    # Simulate file structure
    file_good = Path("/testdir/good.txt")
    file_bad = Path("/testdir/bad.pdf") # This one will cause an error

    configured_mock_files = []
    for p_obj in [file_good, file_bad]:
        m_file = mock.MagicMock(spec=Path)
        m_file.is_file.return_value = True
        m_file.__str__.return_value = str(p_obj)
        configured_mock_files.append(m_file)
    mock_path_rglob.return_value = configured_mock_files

    async def mock_load_document_logic_with_error(filepath_str):
        if filepath_str == str(file_good): return "Good content"
        if filepath_str == str(file_bad):
            # Simulate an exception being returned by asyncio.gather for this task
            raise ValueError("Simulated PDF load error")
        return None

    # Patch DocumentLoader.load_document
    with mock.patch.object(DocumentLoader, 'load_document', side_effect=mock_load_document_logic_with_error) as mock_ld_method:
        caplog.set_level(logging.ERROR) # To catch the "Failed to load document"
        results = await doc_loader.load_from_directory("/testdir")

    assert len(results) == 1
    assert results[0] == (str(file_good), "Good content")

    # Check that the error for bad.pdf was logged by load_from_directory
    assert f"Failed to load document {str(file_bad)} from directory: Simulated PDF load error" in caplog.text

    # Ensure load_document was called for both
    expected_calls = [mock.call(str(file_good)), mock.call(str(file_bad))]
    # Check calls irrespective of order
    assert len(mock_ld_method.call_args_list) == len(expected_calls)
    for c in expected_calls:
        assert c in mock_ld_method.call_args_list, f"Expected call {c} not found in actual calls."
    for c_actual in mock_ld_method.call_args_list:
        assert c_actual in expected_calls, f"Actual call {c_actual} was not expected."

@pytest.mark.asyncio
async def test_load_from_directory_empty_dir(doc_loader, mock_path_is_dir, mock_path_rglob, caplog):
    mock_path_is_dir.return_value = True
    mock_path_rglob.return_value = [] # No files found

    caplog.set_level(logging.INFO)
    results = await doc_loader.load_from_directory("/emptydir")

    assert len(results) == 0
    assert "Loaded 0 documents from directory /emptydir" in caplog.text

# Example of mocking the synchronous internal loaders directly if needed:
@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_pdf_sync_mocked_directly(mock_os_exists, doc_loader, mock_asyncio_to_thread):
    filepath = "test.pdf"
    # Mock the _load_pdf_sync method itself
    with mock.patch.object(DocumentLoader, '_load_pdf_sync', return_value="Directly Mocked PDF Content") as sync_loader_mock:
        content = await doc_loader.load_document(filepath)
        assert content == "Directly Mocked PDF Content"
        mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_pdf_sync, filepath)
        sync_loader_mock.assert_called_once_with(filepath)

# --- Tests that execute synchronous loaders ---

# Helper to create temp files for testing actual loaders
@pytest.fixture
def temp_file_factory(tmp_path):
    def _create_temp_file(filename, content_writer_func, *args):
        file_path = tmp_path / filename
        content_writer_func(file_path, *args)
        return str(file_path)
    return _create_temp_file

# PDF specific writer using PyMuPDF (fitz)
def write_pdf_content(file_path_obj, text_content):
    import fitz # PyMuPDF
    doc = fitz.open() # New empty PDF
    page = doc.new_page()
    page.insert_text((50, 72), text_content)
    doc.save(str(file_path_obj))
    doc.close()

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True) # Assume file exists
async def test_load_document_pdf_actual_loader(mock_os_exists, doc_loader, mock_asyncio_to_thread, temp_file_factory):
    expected_text = "This is a test PDF content."
    # Use the factory to create a real PDF file
    pdf_filepath_str = temp_file_factory("test_actual.pdf", write_pdf_content, expected_text)

    # Ensure _load_pdf_sync is NOT mocked for this test
    # mock_asyncio_to_thread will still ensure it's called in a thread

    content = await doc_loader.load_document(pdf_filepath_str)

    assert content.strip() == expected_text.strip()
    # Check that asyncio.to_thread was called with the real _load_pdf_sync
    # and the correct file path.
    # The passthrough_to_thread mock for asyncio.to_thread directly executes the function,
    # so _load_pdf_sync will have been called with pdf_filepath_str.
    mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_pdf_sync, pdf_filepath_str)

# DOCX specific writer
def write_docx_content(file_path_obj, text_content):
    from docx import Document as DocxDocument # Alias to avoid confusion with DocumentLoader
    doc = DocxDocument()
    doc.add_paragraph(text_content)
    doc.save(str(file_path_obj))

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_docx_actual_loader(mock_os_exists, doc_loader, mock_asyncio_to_thread, temp_file_factory):
    expected_text = "This is test DOCX content."
    docx_filepath_str = temp_file_factory("test_actual.docx", write_docx_content, expected_text)

    content = await doc_loader.load_document(docx_filepath_str)

    assert content.strip() == expected_text.strip()
    mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_docx_sync, docx_filepath_str)

# XLSX specific writer
def write_xlsx_content(file_path_obj, sheet_data_map):
    import openpyxl
    workbook = openpyxl.Workbook()
    # Remove default sheet if creating specific ones
    if "Sheet" in workbook.sheetnames and len(sheet_data_map) > 0 :
        del workbook["Sheet"]

    for sheet_name, rows in sheet_data_map.items():
        sheet = workbook.create_sheet(title=sheet_name)
        for row_data in rows:
            sheet.append(row_data)
    workbook.save(str(file_path_obj))

@pytest.mark.asyncio
@mock.patch("os.path.exists", return_value=True)
async def test_load_document_xlsx_actual_loader(mock_os_exists, doc_loader, mock_asyncio_to_thread, temp_file_factory):
    # Define content for multiple sheets
    sheet1_name = "TestSheet1"
    sheet1_rows = [
        ["Header1", "Header2"],
        ["Data1A", "Data1B"],
        ["Data1C", None], # Test None value
    ]
    sheet2_name = "Another Sheet"
    sheet2_rows = [
        ["Info"],
        ["More data"],
    ]
    xlsx_data = {
        sheet1_name: sheet1_rows,
        sheet2_name: sheet2_rows,
    }

    # Construct expected text based on _load_xlsx_sync logic
    temp_parts_for_expected_text = []
    for sheet_name_key, rows_data in xlsx_data.items():
        temp_parts_for_expected_text.append(f"Sheet: {sheet_name_key}\n")
        for row in rows_data:
            temp_parts_for_expected_text.append(", ".join([str(cell) if cell is not None else "" for cell in row]))
        temp_parts_for_expected_text.append("\n")
    expected_text = "\n".join(temp_parts_for_expected_text)

    xlsx_filepath_str = temp_file_factory("test_actual.xlsx", write_xlsx_content, xlsx_data)

    content = await doc_loader.load_document(xlsx_filepath_str)

    # Normalize newlines and strip for comparison.
    # The actual content might end with \n\n due to the final join, strip() handles this.
    # Expected text also built to match this structure then stripped.
    assert content.replace('\r\n', '\n').strip() == expected_text.strip()
    mock_asyncio_to_thread.assert_called_once_with(DocumentLoader._load_xlsx_sync, xlsx_filepath_str)
